"""
Advanced Excel file handling for trading systems.

This module provides comprehensive Excel operations including:
- Excel file reading and writing
- Worksheet management
- Data formatting and validation
- Trading data import/export
- Chart and table creation
"""

import os
import sys
from typing import List, Optional, Dict, Union, Any, Tuple
from datetime import datetime, date
from dataclasses import dataclass, field
from enum import Enum
import pandas as pd
import numpy as np

# Add path for imports
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))

from src.core.base import CBase, SystemProtocol

# Optional dependencies with graceful fallback
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


class ExcelOperationType(Enum):
    """Types of Excel operations."""
    READ = "read"
    WRITE = "write"
    APPEND = "append"
    CREATE_WORKBOOK = "create_workbook"
    CREATE_WORKSHEET = "create_worksheet"
    DELETE_WORKSHEET = "delete_worksheet"
    FORMAT_CELLS = "format_cells"
    CREATE_CHART = "create_chart"
    EXPORT_DATA = "export_data"
    IMPORT_DATA = "import_data"


class ExcelFormat(Enum):
    """Excel file formats."""
    XLSX = ".xlsx"
    XLS = ".xls"
    XLSM = ".xlsm"
    CSV = ".csv"


@dataclass
class ExcelSheet:
    """Information about an Excel worksheet."""
    
    name: str
    index: int = 0
    row_count: int = 0
    column_count: int = 0
    data_range: str = ""
    headers: List[str] = field(default_factory=list)
    has_data: bool = False


@dataclass
class ExcelWorkbook:
    """Information about an Excel workbook."""
    
    file_path: str
    file_name: str
    file_size: int = 0
    sheets: List[ExcelSheet] = field(default_factory=list)
    creation_date: Optional[datetime] = None
    modification_date: Optional[datetime] = None
    author: str = ""
    is_protected: bool = False


@dataclass
class ExcelOperationResult:
    """Result of an Excel operation."""
    
    operation_type: ExcelOperationType
    success: bool
    file_path: str = ""
    sheet_name: str = ""
    rows_processed: int = 0
    columns_processed: int = 0
    error_message: str = ""
    execution_time_ms: float = 0.0
    data: Any = None


class CExcelFileHandler(CBase):
    """
    Advanced Excel file handler for trading systems.
    
    Provides comprehensive Excel operations including:
    - Reading and writing Excel files (.xlsx, .xls, .xlsm)
    - Worksheet management and manipulation
    - Data formatting and styling
    - Trading data import/export
    - Chart creation for analysis
    """
    
    def __init__(self, system_id: int = 0):
        super().__init__(system_id)
        
        # Configuration
        self.default_sheet_name = "Sheet1"
        self.max_rows_in_memory = 100000
        self.default_date_format = "%Y-%m-%d"
        self.default_datetime_format = "%Y-%m-%d %H:%M:%S"
        
        # Statistics
        self.files_processed = 0
        self.total_rows_processed = 0
        self.total_columns_processed = 0
        
        # Check library availability
        self.openpyxl_available = OPENPYXL_AVAILABLE
        self.xlsxwriter_available = XLSXWRITER_AVAILABLE
        
        if not (self.openpyxl_available or self.xlsxwriter_available):
            print("Warning: Neither openpyxl nor xlsxwriter is available. Excel functionality will be limited.")
    
    def initialize(self, system: SystemProtocol) -> 'CExcelFileHandler':
        """Initialize Excel handler with system."""
        return self
    
    # Basic Excel Operations
    def read_excel_file(self, file_path: str, sheet_name: str = None) -> ExcelOperationResult:
        """Read Excel file and return data."""
        start_time = datetime.now()
        
        if not os.path.exists(file_path):
            return ExcelOperationResult(
                operation_type=ExcelOperationType.READ,
                success=False,
                file_path=file_path,
                error_message="File does not exist"
            )
        
        try:
            # Use pandas for reading (works with multiple engines)
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheets_data = {sheet_name: df}
            else:
                # Read all sheets
                sheets_data = pd.read_excel(file_path, sheet_name=None)
                if isinstance(sheets_data, pd.DataFrame):
                    sheets_data = {self.default_sheet_name: sheets_data}
            
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            
            # Calculate totals
            total_rows = sum(len(df) for df in sheets_data.values())
            total_cols = sum(len(df.columns) for df in sheets_data.values())
            
            self.files_processed += 1
            self.total_rows_processed += total_rows
            self.total_columns_processed += total_cols
            
            return ExcelOperationResult(
                operation_type=ExcelOperationType.READ,
                success=True,
                file_path=file_path,
                sheet_name=sheet_name or "All",
                rows_processed=total_rows,
                columns_processed=total_cols,
                execution_time_ms=execution_time,
                data=sheets_data
            )
            
        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            return ExcelOperationResult(
                operation_type=ExcelOperationType.READ,
                success=False,
                file_path=file_path,
                sheet_name=sheet_name or "",
                error_message=str(e),
                execution_time_ms=execution_time
            )
    
    def write_excel_file(
        self, 
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
        file_path: str,
        sheet_name: str = None,
        index: bool = False,
        header: bool = True
    ) -> ExcelOperationResult:
        """Write data to Excel file."""
        start_time = datetime.now()
        
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            if isinstance(data, pd.DataFrame):
                # Single DataFrame
                sheet_name = sheet_name or self.default_sheet_name
                data_dict = {sheet_name: data}
            else:
                # Multiple DataFrames
                data_dict = data
            
            # Write using pandas ExcelWriter
            with pd.ExcelWriter(file_path, engine='openpyxl' if self.openpyxl_available else 'xlsxwriter') as writer:
                total_rows = 0
                total_cols = 0
                
                for sheet, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet, index=index, header=header)
                    total_rows += len(df)
                    total_cols += len(df.columns)
            
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            
            self.files_processed += 1
            self.total_rows_processed += total_rows
            self.total_columns_processed += total_cols
            
            return ExcelOperationResult(
                operation_type=ExcelOperationType.WRITE,
                success=True,
                file_path=file_path,
                sheet_name=sheet_name or "Multiple",
                rows_processed=total_rows,
                columns_processed=total_cols,
                execution_time_ms=execution_time
            )
            
        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            return ExcelOperationResult(
                operation_type=ExcelOperationType.WRITE,
                success=False,
                file_path=file_path,
                sheet_name=sheet_name or "",
                error_message=str(e),
                execution_time_ms=execution_time
            )
    
    def get_workbook_info(self, file_path: str) -> Optional[ExcelWorkbook]:
        """Get comprehensive workbook information."""
        if not os.path.exists(file_path):
            return None
        
        try:
            # Get file stats
            file_stats = os.stat(file_path)
            file_name = os.path.basename(file_path)
            
            # Read workbook with openpyxl if available
            if self.openpyxl_available:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                
                sheets = []
                for i, sheet_name in enumerate(workbook.sheetnames):
                    ws = workbook[sheet_name]
                    
                    # Get sheet info
                    max_row = ws.max_row
                    max_col = ws.max_column
                    
                    # Get headers if they exist
                    headers = []
                    if max_row > 0:
                        for col in range(1, min(max_col + 1, 50)):  # Limit to first 50 columns
                            cell_value = ws.cell(row=1, column=col).value
                            if cell_value:
                                headers.append(str(cell_value))
                            else:
                                headers.append(f"Column_{col}")
                    
                    sheet_info = ExcelSheet(
                        name=sheet_name,
                        index=i,
                        row_count=max_row,
                        column_count=max_col,
                        data_range=f"A1:{ws.max_column and chr(64 + ws.max_column) or 'A'}{max_row}",
                        headers=headers,
                        has_data=max_row > 1 and max_col > 0
                    )
                    sheets.append(sheet_info)
                
                workbook.close()
                
                return ExcelWorkbook(
                    file_path=file_path,
                    file_name=file_name,
                    file_size=file_stats.st_size,
                    sheets=sheets,
                    creation_date=datetime.fromtimestamp(file_stats.st_ctime),
                    modification_date=datetime.fromtimestamp(file_stats.st_mtime),
                    author="",
                    is_protected=False
                )
            else:
                # Fallback using pandas
                try:
                    sheet_names = pd.ExcelFile(file_path).sheet_names
                    sheets = []
                    
                    for i, sheet_name in enumerate(sheet_names):
                        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1)
                        
                        sheet_info = ExcelSheet(
                            name=sheet_name,
                            index=i,
                            row_count=0,  # Would need full read to get this
                            column_count=len(df.columns),
                            headers=list(df.columns),
                            has_data=len(df.columns) > 0
                        )
                        sheets.append(sheet_info)
                    
                    return ExcelWorkbook(
                        file_path=file_path,
                        file_name=file_name,
                        file_size=file_stats.st_size,
                        sheets=sheets,
                        creation_date=datetime.fromtimestamp(file_stats.st_ctime),
                        modification_date=datetime.fromtimestamp(file_stats.st_mtime)
                    )
                except:
                    return None
                    
        except Exception:
            return None
    
    # Trading Data Specific Operations
    def export_trading_data(
        self,
        trades_data: pd.DataFrame,
        file_path: str,
        include_charts: bool = True,
        format_data: bool = True
    ) -> ExcelOperationResult:
        """Export trading data with formatting and charts."""
        start_time = datetime.now()
        
        try:
            # Prepare data sheets
            sheets_data = {
                "Trades": trades_data.copy()
            }
            
            # Add summary sheet if we have trade data
            if not trades_data.empty and 'pnl' in trades_data.columns:
                summary_data = self._create_trading_summary(trades_data)
                sheets_data["Summary"] = summary_data
            
            # Write base data
            result = self.write_excel_file(sheets_data, file_path)
            
            if result.success and self.openpyxl_available and format_data:
                # Apply formatting
                self._format_trading_workbook(file_path, include_charts)
            
            return result
            
        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            return ExcelOperationResult(
                operation_type=ExcelOperationType.EXPORT_DATA,
                success=False,
                file_path=file_path,
                error_message=str(e),
                execution_time_ms=execution_time
            )
    
    def import_market_data(self, file_path: str, sheet_name: str = None) -> ExcelOperationResult:
        """Import market data from Excel with validation."""
        result = self.read_excel_file(file_path, sheet_name)
        
        if result.success and result.data:
            try:
                # Get the first/specified sheet
                if isinstance(result.data, dict):
                    df = list(result.data.values())[0]
                else:
                    df = result.data
                
                # Validate market data columns
                required_columns = ['date', 'open', 'high', 'low', 'close']
                missing_columns = [col for col in required_columns if col not in df.columns.str.lower()]
                
                if missing_columns:
                    result.error_message = f"Missing required columns: {missing_columns}"
                    result.success = False
                else:
                    # Convert date column
                    date_column = df.columns[df.columns.str.lower() == 'date'][0]
                    df[date_column] = pd.to_datetime(df[date_column])
                    
                    # Ensure numeric columns
                    numeric_columns = ['open', 'high', 'low', 'close', 'volume']
                    for col in numeric_columns:
                        matching_cols = df.columns[df.columns.str.lower() == col]
                        if len(matching_cols) > 0:
                            df[matching_cols[0]] = pd.to_numeric(df[matching_cols[0]], errors='coerce')
                    
                    result.data = df
                    
            except Exception as e:
                result.error_message = f"Data validation failed: {str(e)}"
                result.success = False
        
        return result
    
    # Advanced Operations
    def create_formatted_workbook(self, file_path: str, sheets_config: Dict[str, Dict]) -> ExcelOperationResult:
        """Create a formatted workbook with multiple sheets."""
        if not self.openpyxl_available:
            return ExcelOperationResult(
                operation_type=ExcelOperationType.CREATE_WORKBOOK,
                success=False,
                error_message="openpyxl not available for advanced formatting"
            )
        
        start_time = datetime.now()
        
        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            total_rows = 0
            total_cols = 0
            
            for sheet_name, config in sheets_config.items():
                ws = wb.create_sheet(sheet_name)
                
                # Get data
                data = config.get('data', pd.DataFrame())
                if not data.empty:
                    # Write data
                    for r in dataframe_to_rows(data, index=False, header=True):
                        ws.append(r)
                    
                    total_rows += len(data)
                    total_cols += len(data.columns)
                
                # Apply formatting if specified
                formatting = config.get('formatting', {})
                if formatting:
                    self._apply_sheet_formatting(ws, formatting)
            
            # Save workbook
            wb.save(file_path)
            
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            
            return ExcelOperationResult(
                operation_type=ExcelOperationType.CREATE_WORKBOOK,
                success=True,
                file_path=file_path,
                rows_processed=total_rows,
                columns_processed=total_cols,
                execution_time_ms=execution_time
            )
            
        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds() * 1000
            return ExcelOperationResult(
                operation_type=ExcelOperationType.CREATE_WORKBOOK,
                success=False,
                file_path=file_path,
                error_message=str(e),
                execution_time_ms=execution_time
            )
    
    # Helper Methods
    def _create_trading_summary(self, trades_df: pd.DataFrame) -> pd.DataFrame:
        """Create trading summary from trades data."""
        try:
            summary_data = []
            
            if 'pnl' in trades_df.columns:
                total_trades = len(trades_df)
                winning_trades = len(trades_df[trades_df['pnl'] > 0])
                losing_trades = len(trades_df[trades_df['pnl'] < 0])
                
                total_pnl = trades_df['pnl'].sum()
                avg_win = trades_df[trades_df['pnl'] > 0]['pnl'].mean() if winning_trades > 0 else 0
                avg_loss = trades_df[trades_df['pnl'] < 0]['pnl'].mean() if losing_trades > 0 else 0
                
                summary_data = [
                    ['Total Trades', total_trades],
                    ['Winning Trades', winning_trades],
                    ['Losing Trades', losing_trades],
                    ['Win Rate %', (winning_trades / total_trades * 100) if total_trades > 0 else 0],
                    ['Total P&L', total_pnl],
                    ['Average Win', avg_win],
                    ['Average Loss', avg_loss],
                    ['Profit Factor', (avg_win * winning_trades) / abs(avg_loss * losing_trades) if avg_loss != 0 else 0],
                ]
            
            return pd.DataFrame(summary_data, columns=['Metric', 'Value'])
            
        except Exception:
            return pd.DataFrame()
    
    def _format_trading_workbook(self, file_path: str, include_charts: bool = True):
        """Apply formatting to trading workbook."""
        if not self.openpyxl_available:
            return
        
        try:
            wb = load_workbook(file_path)
            
            # Format trades sheet
            if "Trades" in wb.sheetnames:
                ws_trades = wb["Trades"]
                self._format_trades_sheet(ws_trades)
            
            # Format summary sheet  
            if "Summary" in wb.sheetnames:
                ws_summary = wb["Summary"]
                self._format_summary_sheet(ws_summary)
                
                if include_charts:
                    self._add_trading_chart(ws_summary)
            
            wb.save(file_path)
            
        except Exception:
            pass  # Fail silently on formatting errors
    
    def _format_trades_sheet(self, worksheet):
        """Format the trades worksheet."""
        # Header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_summary_sheet(self, worksheet):
        """Format the summary worksheet."""
        # Header formatting for summary
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['B1'].font = Font(bold=True, size=14)
        
        # Value formatting
        for row in range(2, worksheet.max_row + 1):
            if worksheet[f'A{row}'].value and 'P&L' in str(worksheet[f'A{row}'].value):
                worksheet[f'B{row}'].number_format = '#,##0.00'
            elif worksheet[f'A{row}'].value and '%' in str(worksheet[f'A{row}'].value):
                worksheet[f'B{row}'].number_format = '0.00%'
    
    def _add_trading_chart(self, worksheet):
        """Add a simple chart to the summary sheet."""
        try:
            # This is a simplified chart - would need actual P&L data points
            chart = LineChart()
            chart.title = "Trading Performance"
            chart.style = 13
            chart.y_axis.title = 'P&L'
            chart.x_axis.title = 'Trades'
            
            # Would need actual data range here
            # data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row)
            # chart.add_data(data, titles_from_data=True)
            
            # worksheet.add_chart(chart, "D2")
            
        except Exception:
            pass  # Chart creation is optional
    
    def _apply_sheet_formatting(self, worksheet, formatting_config: Dict):
        """Apply custom formatting to worksheet."""
        # Header styling
        if 'header_style' in formatting_config:
            header_style = formatting_config['header_style']
            for cell in worksheet[1]:
                if 'font' in header_style:
                    cell.font = Font(**header_style['font'])
                if 'fill' in header_style:
                    cell.fill = PatternFill(**header_style['fill'])
        
        # Column widths
        if 'column_widths' in formatting_config:
            for col_letter, width in formatting_config['column_widths'].items():
                worksheet.column_dimensions[col_letter].width = width
    
    # Statistics and Information
    def get_operation_statistics(self) -> Dict[str, Any]:
        """Get Excel operation statistics."""
        return {
            "files_processed": self.files_processed,
            "total_rows_processed": self.total_rows_processed,
            "total_columns_processed": self.total_columns_processed,
            "openpyxl_available": self.openpyxl_available,
            "xlsxwriter_available": self.xlsxwriter_available,
            "default_sheet_name": self.default_sheet_name
        }
    
    def __str__(self) -> str:
        """String representation of Excel handler."""
        return (
            f"CExcelFileHandler(files_processed={self.files_processed}, "
            f"rows_processed={self.total_rows_processed}, "
            f"openpyxl={'✓' if self.openpyxl_available else '✗'})"
        )